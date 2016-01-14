from django.http import HttpResponse
from django.template import loader, RequestContext
from django import forms
from os import path, chdir
import itunesreviewreport
import io
from collections import OrderedDict
from openpyxl import Workbook

from django.views.generic import TemplateView

def main(request,ID = ""):

    try:
        ID = request.POST['ID']
    except:
        ID = ""

    countryfile=0;

    #get the country list
    #run through the country list file and generate country objects with the country codes and names
    try:
        newPath = path.abspath(path.dirname(__file__))
        chdir(newPath)
        
    except Exception, e:
        return HttpResponse("INTERNAL ERROR: countryList not opened, file missing from app folder")

    try:
        countryfile = open("countrylist.csv").readlines()
        #parse countryfile into dictionary {Region:[country1,country2...],...}
        countryDict = {}
        for c in countryfile:
            sp = c.split(',')
            try:
                countryDict[sp[2]].append(sp[0]+","+sp[1])
            except:
                countryDict[sp[2]] = [sp[0]+","+sp[1]]

        #return HttpResponse(sorted(countryDict.items()))

        sortedCountryDict = OrderedDict(sorted(countryDict.items()))

    except:
        return HttpResponse("can't open countryList, file missing from app folder")
    
    template = loader.get_template('itunesreviews/main.html')
    context = RequestContext(request, {'country_list' : sortedCountryDict, 'Ident':ID,})
    return HttpResponse(template.render(context))
     
def runReport(request):
    try:
        idnumber = request.POST['id']
        countryList = request.POST.getlist('countryList[]')
        #nextCountry = request.POST['nextCountry'] # should be an int to pass to countryList
    

        #validate input and get report
        if(int(idnumber) < 10000000000 and int(idnumber) >= 100000000):
            template = loader.get_template('itunesreviews/progress.html')
            #create the workbook
            book = Workbook() 
            #create the summary page
            summarySheet = book.create_sheet()
            summarySheet.title = "Summary"    
            summarySheet.cell(row=1,column=1).value = "Country"
            summarySheet.cell(row=1,column=2).value = "Reviews"
            #remove the default sheet
            book.remove_sheet(book.get_sheet_by_name("Sheet"))
            
            report = itunesreviewreport.report(idnumber,countryList,book,summarySheet)
            
            response = HttpResponse(report ,content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="iTunesReviews.xls"'
            
            return(response)
        else:
            #pass a list with the next country to do
            return HttpResponse("itunes id error, " +  idnumber +  " not in range")
            
    except:
        return HttpResponse(request.POST['countryList[]'])

def progress(request):

    try:
        idnumber = request.POST['id']
        countryList = request.POST.getlist('countryList[])')
        #nextCountry = request.POST['nextCountry'] # should be an int to pass to countryList
    except:
        return HttpResponse("Internal Error")

    template = loader.get_template('itunesreviews/progress.html')

    #create the workbook
    book = Workbook() 
    #create the summary page
    summarySheet = book.create_sheet()
    summarySheet.title = "Summary"    
    summarySheet.cell(row=1,column=1).value = "Country"
    summarySheet.cell(row=1,column=2).value = "Reviews"
    #remove the default sheet
    book.remove_sheet(book.get_sheet_by_name("Sheet"))

    report = itunesreviewreport.report(idnumber,countryList,book,summarySheet)

    response = HttpResponse(report ,content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="iTunesReviews.xls"'

    return(report);

    #pass a list with the next country to do

def search(request):
    try:
        name = request.POST['name']
    except:
        name = ""

    template = loader.get_template('itunesreviews/search.html')

    #if there is a search term do the search, otherwise render the empty results page
    if(name != None and name != ""):
        context = RequestContext(request, {'results' : itunesreviewreport.idLookup(name), 'SearchTerm':name,})
        return HttpResponse(template.render(context))
    else:
        context = RequestContext(request, {'SearchTerm':name,})
        return HttpResponse(template.render(context))

#itunesreviewreport.report(903688996)
