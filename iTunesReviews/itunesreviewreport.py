# -*- coding: utf-8 -*-
import xml.etree.ElementTree as ET
import requests
from os import path, chdir
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

class review:
    def __init__(self, title="", userName="", date="", text="", rating = 0):
        self.title = title
        self.userName = userName
        self.date = date
        self.text = text
        self.rating = rating

class country:
    #class to hold all the data from one country
    def __init__(self, code, name, reviewList, iTunesID):
        self.code = code
        self.name = name
        self.reviews = reviewList
        self.iTunesID = iTunesID
         
    def getReviews(self, page = 1, recursion = True): 
    #takes a country object and makes a call to itunes to gather all review data from that store.
    #adds review data to country objects

        
        url = "https://itunes.apple.com/" + self.code + "/rss/customerreviews" + "/page=" + str(page) + "/id=" + str(self.iTunesID) +  "/sortBy=mostRecent/xml"

        resp = requests.get(url, verify = "cacert.pem")
        resp = resp.text
        resp = resp.encode("utf8")
        
        tree = ET.fromstring(resp)
        
        #determine if there are more xml pages to view
        lastPageNum = None
        st = "{http://www.w3.org/2005/Atom}" #standard text that is before all tags
        
        #loop to get the reviews from the xml
        for e in tree:
            if e.tag == st+"entry" and e != tree[12]:
                self.reviews.append(review(e.find(st+'title').text, e.find(st+'author').find(st+'name').text,e.find(st+'updated').text,e.find(st+'content').text ,e.find('{http://itunes.apple.com/rss}rating').text))
                
            if e.tag == st+"link":
                if e.get('rel') == "last" and e.get("href")!="":
                    lastPage = e.get("href")
                    temp = lastPage.rsplit('page=')
                    lastPageNum = int(temp[1][:1])
                elif e.get("href")=="":
                    lastPageNum = page

        if lastPageNum > page and recursion == True:
            page += 1        
            self.getReviews(page,True)
   
                
def report(itunesID):
    """maskes a list of country objects and generates an excel sheet with all the reviews from all itunes stores"""
    failedSearches = []
    
    #run through the country list file and generate country objects with the country codes and names
    try:
        newPath = path.abspath(path.dirname(__file__))
        chdir(newPath)
    
    except Exception, e:
        return str(e)
    
    try:
        countryfile = open("countrylist.csv").readlines()
        countryList = []
    except:
        return "can't open countryList, please move to the same folder as the program"
    
    
    for c in countryfile:
        newCountryObject = country(c[0:2],c[3:-1],[],itunesID)
        countryList.append(newCountryObject)
        
        try:
            newCountryObject.getReviews() #while looping through collect review data        
        except Exception, e:
            print e
            if str(e) != "no element found: line 1, column 0": #surpressing this error to simply skip these countries instead         
                return str(e) + " - " + newCountryObject.name
            else:
                failedSearches.append(newCountryObject.name)
            
    
    book = Workbook() 
    
    outputPath = newPath + "\itunesReviews.xls"
    
    #create the summary page
    summarySheet = book.create_sheet()
    summarySheet.title = "Summary"
    
    summarySheet.cell(row=1,column=1).value = "Country"
    summarySheet.cell(row=1,column=2).value = "Reviews"
        
    book.remove_sheet(book.get_sheet_by_name("Sheet"))#remove the default sheet
    workingRow = 2

    noReviews = True
    summaryList = []
    
    for con in countryList:
                
        summaryList.append((con.name,len(con.reviews)))    
        
        if(len(con.reviews) > 0):
               
            
            if noReviews == True:
                noReviews = False
                
            workingSheet = book.create_sheet(title = con.name)
            revRow = 2
            
            workingSheet.cell(row=1,column=1).value = "Title"
            workingSheet.cell(row=1,column=2).value = "Date"
            workingSheet.cell(row=1,column=3).value = "UserName"
            workingSheet.cell(row=1,column=4).value = "Text"
            workingSheet.cell(row=1,column=5).value = "Rating"
            
            for rev in con.reviews: 
                workingSheet.cell(row=revRow,column=1).value = rev.title
                workingSheet.cell(row=revRow,column=2).value = rev.date
                workingSheet.cell(row=revRow,column=3).value = rev.userName
                workingSheet.cell(row=revRow,column=4).value = rev.text
                workingSheet.cell(row=revRow,column=5).value = rev.rating
                
                revRow += 1
    
        workingRow+=1
    
    workingRow = 1
    summaryList.sort(key=lambda tup: tup[1], reverse = True)
        
    for k in summaryList: 
        summarySheet.cell(row=workingRow,column=1).value = k[0]
        summarySheet.cell(row=workingRow,column=2).value = k[1]
        workingRow += 1


    if noReviews == True:
    
        return "No reviews found. If you believe this is incorrect please check your iTunes ID"
    
    book.save(outputPath) 
    return save_virtual_workbook(book)
    
    if len(failedSearches) > 0:
        print failedSearches
        return "Complete, failed searches from:" + str(failedSearches)
    


    
    
    